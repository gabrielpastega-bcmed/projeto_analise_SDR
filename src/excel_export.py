"""
Excel export utilities with professional formatting.

Creates multi-sheet Excel files with formatted data and charts.
"""

from datetime import datetime
from io import BytesIO
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet


class ExcelExporter:
    """Excel exporter with professional formatting."""

    # Color scheme
    HEADER_BG = "1F4E78"  # Dark blue
    HEADER_FG = "FFFFFF"  # White
    ALT_ROW_BG = "F2F2F2"  # Light gray
    BORDER_COLOR = "D3D3D3"  # Light gray border

    def __init__(self):
        """Initialize Excel exporter."""
        self.workbook = Workbook()
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

    def _format_header(self, ws: Worksheet, end_column: int) -> None:
        """
        Apply header formatting to first row.

        Args:
            ws: Worksheet to format
            end_column: Last column index (1-based)
        """
        # Header style
        header_font = Font(bold=True, color=self.HEADER_FG, size=11)
        header_fill = PatternFill(start_color=self.HEADER_BG, end_color=self.HEADER_BG, fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Apply to header row
        for col in range(1, end_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _apply_borders(self, ws: Worksheet, max_row: int, max_col: int) -> None:
        """
        Apply borders to all cells.

        Args:
            ws: Worksheet
            max_row: Maximum row
            max_col: Maximum column
        """
        thin_border = Border(
            left=Side(style="thin", color=self.BORDER_COLOR),
            right=Side(style="thin", color=self.BORDER_COLOR),
            top=Side(style="thin", color=self.BORDER_COLOR),
            bottom=Side(style="thin", color=self.BORDER_COLOR),
        )

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = thin_border

    def _apply_alternating_rows(self, ws: Worksheet, max_row: int, max_col: int) -> None:
        """
        Apply alternating row colors (zebra stripes).

        Args:
            ws: Worksheet
            max_row: Maximum row
            max_col: Maximum column
        """
        alt_fill = PatternFill(start_color=self.ALT_ROW_BG, end_color=self.ALT_ROW_BG, fill_type="solid")

        # Start from row 2 (skip header)
        for row in range(2, max_row + 1):
            if row % 2 == 0:
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = alt_fill

    def _auto_adjust_columns(self, ws: Worksheet) -> None:
        """
        Auto-adjust column widths based on content.

        Args:
            ws: Worksheet
        """
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def add_dataframe_sheet(self, df: pd.DataFrame, sheet_name: str, freeze_header: bool = True) -> Worksheet:
        """
        Add a formatted dataframe as a new sheet.

        Args:
            df: DataFrame to add
            sheet_name: Name of the sheet
            freeze_header: Whether to freeze first row

        Returns:
            Created worksheet
        """
        ws = self.workbook.create_sheet(title=sheet_name)

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Apply formatting
        self._format_header(ws, len(df.columns))
        self._apply_borders(ws, len(df) + 1, len(df.columns))
        self._apply_alternating_rows(ws, len(df) + 1, len(df.columns))
        self._auto_adjust_columns(ws)

        # Freeze header row
        if freeze_header:
            ws.freeze_panes = "A2"

        return ws

    def add_summary_sheet(self, summary_data: dict[str, Any], sheet_name: str = "Resumo") -> Worksheet:
        """
        Add a summary sheet with key-value pairs.

        Args:
            summary_data: Dictionary with summary metrics
            sheet_name: Name of the sheet

        Returns:
            Created worksheet
        """
        ws = self.workbook.create_sheet(title=sheet_name)

        # Title
        ws["A1"] = "Métrica"
        ws["B1"] = "Valor"

        # Add data
        for idx, (key, value) in enumerate(summary_data.items(), start=2):
            ws[f"A{idx}"] = key
            ws[f"B{idx}"] = value

        # Format
        self._format_header(ws, 2)
        self._apply_borders(ws, len(summary_data) + 1, 2)
        self._auto_adjust_columns(ws)

        return ws

    def add_chart(
        self,
        ws: Worksheet,
        chart_type: str,
        data_range: str,
        title: str,
        position: str = "D2",
    ) -> None:
        """
        Add a chart to worksheet.

        Args:
            ws: Worksheet
            chart_type: 'bar' or 'pie'
            data_range: Excel range for data (e.g., "A1:B5")
            title: Chart title
            position: Position to place chart (e.g., "D2")
        """
        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            return

        chart.title = title
        chart.style = 10

        # Parse range (simplified)
        # For production, use proper range parsing
        data = Reference(ws, min_col=1, min_row=1, max_col=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)

        ws.add_chart(chart, position)

    def save_to_bytes(self) -> BytesIO:
        """
        Save workbook to BytesIO buffer.

        Returns:
            BytesIO buffer with Excel file
        """
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def save_to_file(self, filename: str) -> None:
        """
        Save workbook to file.

        Args:
            filename: Path to save file
        """
        self.workbook.save(filename)


def create_chat_export(chats: list, filters: Optional[dict] = None) -> BytesIO:
    """
    Create Excel export from chat data.

    Args:
        chats: List of chat objects
        filters: Optional filter information

    Returns:
        BytesIO buffer with Excel file
    """
    from src.dashboard_utils import (
        classify_lead_qualification,
        get_chat_tags,
        get_lead_origin,
    )

    exporter = ExcelExporter()

    # Create summary data
    summary = {
        "Data de Exportação": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "Total de Atendimentos": len(chats),
        "Filtros Ativos": str(filters) if filters else "Nenhum",
    }

    exporter.add_summary_sheet(summary)

    # Create detailed data
    data = []
    for chat in chats:
        data.append(
            {
                "ID": chat.id,
                "Data": (chat.timestamp.strftime("%d/%m/%Y %H:%M") if chat.timestamp else "N/A"),
                "Agente": chat.agentName or "N/A",
                "Origem": get_lead_origin(chat),
                "TME (min)": (f"{chat.waitingTime / 60:.1f}" if chat.waitingTime else "0"),
                "Mensagens": chat.messagesCount or len(chat.messages),
                "Qualificação": classify_lead_qualification(get_chat_tags(chat)),
                "Tags": ", ".join(get_chat_tags(chat)),
            }
        )

    df = pd.DataFrame(data)
    exporter.add_dataframe_sheet(df, "Atendimentos Detalhados")

    # Create agent summary
    agent_stats = {}
    for chat in chats:
        agent = chat.agentName or "Desconhecido"
        if agent not in agent_stats:
            agent_stats[agent] = {"count": 0, "total_tme": 0}

        agent_stats[agent]["count"] += 1
        if chat.waitingTime:
            agent_stats[agent]["total_tme"] += chat.waitingTime / 60

    agent_df = pd.DataFrame(
        [
            {
                "Agente": agent,
                "Atendimentos": stats["count"],
                "TME Médio (min)": (f"{stats['total_tme'] / stats['count']:.1f}" if stats["count"] > 0 else "0"),
            }
            for agent, stats in agent_stats.items()
        ]
    )
    exporter.add_dataframe_sheet(agent_df, "Resumo por Agente")

    return exporter.save_to_bytes()
